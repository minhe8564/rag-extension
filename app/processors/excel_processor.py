"""
Excel 프로세서
Excel 파일(.xlsx, .xls)을 PDF로 변환 후 Marker로 처리
행과 열을 전치(transpose)하여 읽음 (질문1 답변1 순서로)
"""
import logging
import threading
import torch
import tempfile
from pathlib import Path
from typing import Dict, Any, List

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from marker.converters.pdf import PdfConverter
from marker.models import create_model_dict
from marker.output import text_from_rendered

from app.core.settings import settings
from app.core.utils.pdf_converter import convert_to_pdf_and_upload
from .base import BaseProcessor

logger = logging.getLogger(__name__)


class ExcelProcessor(BaseProcessor):
    """
    Excel 파일(.xlsx, .xls)을 Markdown으로 변환하는 프로세서
    PDF로 변환 후 Marker로 처리 (이미지 처리를 위해)
    """
    
    @property
    def supported_extensions(self) -> List[str]:
        """
        지원하는 파일 확장자
        """
        return [".xlsx", ".xls"]
    
    def __init__(self):
        self._model_dict = None
        self._lock = threading.Lock()
        
        # settings에서 명시적으로 device가 지정된 경우
        requested_device = settings.marker_device.lower()
        if requested_device in ["cuda", "mps", "cpu"]:
            if requested_device == "cuda" and torch.cuda.is_available():
                self._device = "cuda"
            elif requested_device == "mps" and torch.backends.mps.is_available():
                self._device = "mps"
            elif requested_device == "cpu":
                self._device = "cpu"
            else:
                logger.warning(f"{requested_device}를 사용하도록 설정되어 있지만 사용할 수 없습니다. 자동으로 다른 device로 전환합니다.")
                self._device = self._auto_select_device()
        else:
            self._device = self._auto_select_device()
        
        # dtype 설정
        if self._device == "cpu":
            self._dtype = "float32"
            self._dtype_t = torch.float32
        else:
            self._dtype = settings.marker_dtype.lower() if settings.marker_dtype else "float16"
            self._dtype_t = getattr(torch, self._dtype, torch.float16)
        
    def _auto_select_device(self) -> str:
        """
        자동으로 device 선택 (우선순위: CUDA > MPS > CPU)
        """
        if torch.cuda.is_available():
            return "cuda"
        elif torch.backends.mps.is_available():
            return "mps"
        else:
            return "cpu"
    
    def _ensure_model(self):
        """
        Marker 모델을 싱글톤으로 로드 (thread-safe)
        PDFProcessor와 동일한 모델 공유
        """
        if self._model_dict is not None:
            return self._model_dict
        
        with self._lock:
            if self._model_dict is not None:
                return self._model_dict
            
            try:
                logger.info(f"[Excel Marker 모델] 로딩 시작 - Device: {self._device.upper()}, Dtype: {self._dtype}")
                kwargs = {"device": self._device}
                if self._dtype_t is not None:
                    kwargs["dtype"] = self._dtype_t
                self._model_dict = create_model_dict(**kwargs)
                logger.info(f"[Excel Marker 모델] 로드 완료 - Device: {self._device.upper()}, Dtype: {self._dtype}")
            except TypeError:
                logger.info(f"[Excel Marker 모델] 로딩 시작 (dtype 제외) - Device: {self._device.upper()}")
                self._model_dict = create_model_dict(device=self._device)
                logger.info(f"[Excel Marker 모델] 로드 완료 - Device: {self._device.upper()}")
            
            return self._model_dict
    
    def process(self, file_path: str) -> Dict[str, Any]:
        """
        Excel 파일 -> 행/열 전치 -> PDF 변환 -> Marker로 Markdown 변환
        행과 열을 전치하여 질문1 답변1 순서로 읽음
        """
        if not Path(file_path).exists():
            raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")
        
        pdf_path = None
        minio_path = None
        transposed_excel_path = None
        try:
            # 1. Excel 파일을 읽어서 행과 열을 전치(transpose)
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext == ".xlsx" and PANDAS_AVAILABLE:
                # pandas로 읽어서 전치
                logger.info(f"Excel 파일 읽기 및 전치 시작: {file_path}")
                transposed_excel_path = self._transpose_excel_with_pandas(file_path)
                excel_to_convert = transposed_excel_path
            elif file_ext == ".xlsx" and OPENPYXL_AVAILABLE:
                # openpyxl로 읽어서 전치
                logger.info(f"Excel 파일 읽기 및 전치 시작: {file_path}")
                transposed_excel_path = self._transpose_excel_with_openpyxl(file_path)
                excel_to_convert = transposed_excel_path
            else:
                # .xls 파일이거나 pandas/openpyxl이 없는 경우 원본 사용
                logger.info(f"Excel 파일 전치 건너뛰기 (원본 사용): {file_path}")
                excel_to_convert = file_path
            
            # 2. 전치된 Excel 파일을 PDF로 변환하고 MinIO에 업로드
            # 원본 파일 경로를 전달하여 이미지 추출 시 원본 파일 사용
            logger.info(f"Excel 파일을 PDF로 변환 시작: {excel_to_convert}")
            original_path = file_path if transposed_excel_path else None
            pdf_path, minio_path = convert_to_pdf_and_upload(excel_to_convert, original_file_path=original_path)
            logger.info(f"PDF 변환 완료: {pdf_path}")
            if minio_path:
                logger.info(f"MinIO 업로드 완료: {minio_path}")
            
            # 3. PDF를 Marker로 처리
            mdict = self._ensure_model()
            conv = PdfConverter(artifact_dict=mdict)
            rendered = conv(pdf_path)
            md_text, _, images = text_from_rendered(rendered)
            
            logger.info(f"Excel 파일 처리 완료: {file_path}")
            
            return {
                "content": md_text,
                "metadata": {
                    "file_type": Path(file_path).suffix.lower(),
                    "device": self._device,
                    "dtype": self._dtype,
                    "image_count": len(images) if images else 0,
                    "transposed": transposed_excel_path is not None,
                    "minio_path": minio_path,  # MinIO 업로드 경로 추가
                }
            }
        except Exception as e:
            logger.error(f"Excel 처리 실패: {file_path}, 오류: {e}", exc_info=True)
            raise
        finally:
            # 임시 파일 정리
            if pdf_path and Path(pdf_path).exists():
                try:
                    if tempfile.gettempdir() in str(pdf_path):
                        Path(pdf_path).unlink()
                        logger.debug(f"임시 PDF 파일 삭제: {pdf_path}")
                except Exception as e:
                    logger.warning(f"임시 PDF 파일 삭제 실패: {pdf_path}, 오류: {e}")
            
            if transposed_excel_path and Path(transposed_excel_path).exists():
                try:
                    if tempfile.gettempdir() in str(transposed_excel_path):
                        Path(transposed_excel_path).unlink()
                        logger.debug(f"임시 전치 Excel 파일 삭제: {transposed_excel_path}")
                except Exception as e:
                    logger.warning(f"임시 전치 Excel 파일 삭제 실패: {transposed_excel_path}, 오류: {e}")
    
    def _transpose_excel_with_pandas(self, file_path: str) -> str:
        """
        pandas를 사용하여 Excel 파일을 읽고 행/열을 전치한 후 임시 파일로 저장
        """
        try:
            # 모든 시트 읽기
            excel_file = pd.ExcelFile(file_path)
            
            # 임시 파일 생성
            temp_dir = Path(tempfile.gettempdir())
            temp_file = temp_dir / f"transposed_{Path(file_path).stem}_{Path(file_path).suffix}"
            
            with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
                for sheet_name in excel_file.sheet_names:
                    # 시트 읽기
                    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                    
                    # 행과 열 전치
                    df_transposed = df.T
                    
                    # 전치된 데이터를 새 시트에 저장
                    df_transposed.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            
            logger.info(f"Excel 파일 전치 완료 (pandas): {file_path} -> {temp_file}")
            return str(temp_file)
        except Exception as e:
            logger.error(f"pandas로 Excel 전치 실패: {e}")
            raise
    
    def _transpose_excel_with_openpyxl(self, file_path: str) -> str:
        """
        openpyxl을 사용하여 Excel 파일을 읽고 행/열을 전치한 후 임시 파일로 저장
        """
        try:
            # Excel 파일 열기
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # 새 워크북 생성
            new_wb = openpyxl.Workbook()
            new_wb.remove(new_wb.active)  # 기본 시트 제거
            
            # 각 시트 처리
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 모든 셀 데이터 읽기
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(list(row))
                
                # 행과 열 전치
                if data:
                    transposed_data = list(map(list, zip(*data)))
                else:
                    transposed_data = []
                
                # 새 시트 생성
                new_ws = new_wb.create_sheet(title=sheet_name)
                
                # 전치된 데이터 쓰기
                for row_idx, row_data in enumerate(transposed_data, start=1):
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        new_ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # 임시 파일로 저장
            temp_dir = Path(tempfile.gettempdir())
            temp_file = temp_dir / f"transposed_{Path(file_path).stem}_{Path(file_path).suffix}"
            new_wb.save(temp_file)
            new_wb.close()
            wb.close()
            
            logger.info(f"Excel 파일 전치 완료 (openpyxl): {file_path} -> {temp_file}")
            return str(temp_file)
        except Exception as e:
            logger.error(f"openpyxl로 Excel 전치 실패: {e}")
            raise