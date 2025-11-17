from .base import BaseExtractionStrategy
from typing import Dict, Any, List
from loguru import logger
import os
from app.service.minio_client import ensure_bucket, put_object_bytes

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None
    logger.warning("openpyxl not installed. XLSX extraction will not work.")


class Openpyxl(BaseExtractionStrategy):
    """Basic XLSX 추출 전략 - openpyxl 사용"""
    
    def extract(self, file_path: str) -> Dict[Any, Any]:
        """XLSX 파일 추출 처리"""
        logger.info(f"[BasicXlsx] Extracting XLSX file: {file_path}")

        if load_workbook is None:
            raise ImportError("openpyxl is required for XLSX extraction. Install it with: pip install openpyxl")

        try:
            # XLSX 파일 열기
            workbook = load_workbook(file_path, data_only=True)
            
            # 모든 시트 데이터 추출
            sheets_data = {}
            sheets_text = []

            # 진행률 콜백 (옵션)
            progress_cb = None
            try:
                progress_cb = self.parameters.get("progress_cb") if isinstance(self.parameters, dict) else None
            except Exception:
                progress_cb = None

            # 전체 행 수(대략)를 합산해 total로 사용
            total_units = 0
            try:
                for sheet_name in workbook.sheetnames:
                    try:
                        total_units += workbook[sheet_name].max_row or 0
                    except Exception:
                        pass
            except Exception:
                total_units = 0
            processed = 0
            if progress_cb and total_units:
                try:
                    progress_cb(processed, total_units)
                except Exception:
                    pass

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # 시트의 모든 셀 데이터 추출
                sheet_rows = []
                sheet_text_rows = []

                for row in sheet.iter_rows(values_only=True):
                    # None 값을 빈 문자열로 변환
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    sheet_rows.append(row_data)

                    # 텍스트로 합치기 (탭으로 구분)
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    sheet_text_rows.append(row_text)

                    # 진행률 업데이트: 행 단위로 1 증가
                    processed += 1
                    if progress_cb and total_units:
                        try:
                            progress_cb(processed, total_units)
                        except Exception:
                            pass

                sheets_data[sheet_name] = {
                    "rows": sheet_rows,
                    "row_count": len(sheet_rows)
                }
                
                # 시트별 텍스트 합치기
                sheet_text = "\n".join(sheet_text_rows)
                sheets_text.append({
                    "sheet_name": sheet_name,
                    "content": sheet_text
                })
            
            workbook.close()
            
            # 전체 텍스트 합치기
            full_text = "\n\n".join([sheet["content"] for sheet in sheets_text])
            
            # MinIO 업로드
            try:
                user_id = (self.parameters.get("user_id") or "unknown-user") if isinstance(self.parameters, dict) else "unknown-user"
                file_name = (self.parameters.get("file_name") or "extracted.xlsx") if isinstance(self.parameters, dict) else "extracted.xlsx"
                base_name = os.path.splitext(file_name)[0] or "extracted"
                object_name = f"{user_id}/{base_name}.txt"
                bucket = "ingest"
                ensure_bucket(bucket)
                put_object_bytes(bucket, object_name, full_text.encode("utf-8"), content_type="text/plain; charset=utf-8")
                return {"full_text": full_text, "bucket": bucket, "path": object_name}
            except Exception as e:
                logger.warning(f"[BasicXlsx] MinIO upload failed: {e}")
                return {"full_text": full_text}
        except Exception as e:
            logger.error(f"[BasicXlsx] Error extracting XLSX: {str(e)}")
            raise

